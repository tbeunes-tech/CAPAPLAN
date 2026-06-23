"""Migration des données existantes (§9).

Lit le classeur `PORTFOLIO PROJETS DSI V_4_02.xlsm` et alimente les 4 tables.

Usage::

    python -m scripts.migrate_from_xlsm "/chemin/PORTFOLIO PROJETS DSI V_4_02.xlsm"
    python -m scripts.migrate_from_xlsm <xlsm> --database-url sqlite:///./portfolio.db

Le script est **idempotent** (upsert par clé) : on peut le rejouer sans dupliquer.

Décisions (cf. AGENTS.md) :
- `projects.in_plan` est **recalculé** depuis le statut (§6.2), pas recopié de la colonne « IN PLAN ? ».
- `monthly_loads.in_plan` reprend la valeur **stockée** du classeur (snapshot fidèle) pour que les
  dashboards recalculés correspondent à ceux du classeur (test de non-régression §9).
- `total_project_load` et `last_update` sont **recalculés** après chargement des charges.
- Équipes référencées dans les charges/capacités mais absentes de l'onglet `Equipes` :
  créées en **stub** et listées dans le rapport (aucune donnée perdue, FK respectées).
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import enums
from app.config import settings
from app.database import Base
from app.models import MonthlyCapacity, MonthlyLoad, Project, Team
from app.services.rules import is_in_plan

# Champs projet à liste fermée (§4) → on signale les valeurs hors référentiel (sans bloquer).
_REFERENTIAL_FIELDS = {
    "entite": enums.ENTITE,
    "domain_lead": enums.DOMAIN_LEAD,
    "status": enums.STATUT,
    "priorite": enums.PRIORITE,
    "pilier_strategique": enums.PILIER_STRATEGIQUE,
    "programme": enums.PROGRAMME,
}

SHEET_PROJECTS = "Project Portfolio 25-26"
SHEET_LOADS = "Détail Mensuel"
SHEET_TEAMS = "Equipes"
SHEET_CAPA = "Détail Capa Mensuelle"

TRUE_TOKENS = {"VRAI", "TRUE", "OUI", "YES", "1", "X", "Y"}


# --------------------------------------------------------------------------- #
# Normalisation
# --------------------------------------------------------------------------- #
def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _to_dt(v) -> datetime | None:
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return None


def _first_of_month(v) -> date | None:
    d = _to_date(v)
    return date(d.year, d.month, 1) if d else None


def _norm_bool(v) -> bool:
    """Normalise InPlan (`VRAI/TRUE/OUI/YES` → true), tolère booléens et nombres."""
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return v != 0
    return str(v).strip().upper() in TRUE_TOKENS


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
class Report:
    def __init__(self):
        self.teams = 0
        self.stub_teams: list[str] = []
        self.projects = 0
        self.loads = 0
        self.capacity = 0
        self.skipped_loads = 0
        self.skipped_capa = 0
        self.warnings: list[str] = []

    def render(self) -> str:
        lines = [
            "── Rapport de migration ──────────────────────────────",
            f"  teams importées        : {self.teams}",
            f"  projets importés       : {self.projects}",
            f"  charges (monthly_loads): {self.loads}  (ignorées: {self.skipped_loads})",
            f"  capacités              : {self.capacity}  (ignorées: {self.skipped_capa})",
        ]
        if self.stub_teams:
            lines.append(
                f"  ⚠ {len(self.stub_teams)} équipe(s) STUB créée(s) (référencées mais absentes de "
                f"l'onglet « {SHEET_TEAMS} ») : {', '.join(sorted(self.stub_teams))}"
            )
        for w in self.warnings:
            lines.append(f"  ⚠ {w}")
        lines.append("──────────────────────────────────────────────────────")
        return "\n".join(lines)


def migrate(xlsm_path: str, session) -> Report:
    # L'import massif initial n'est pas journalisé (§8.2 vise les écritures applicatives).
    # On pose le drapeau le temps de l'import puis on le réinitialise (évite toute fuite de
    # contexte vers d'autres sessions/tests).
    from app.audit import audit_disabled
    token = audit_disabled.set(True)
    try:
        return _migrate(xlsm_path, session)
    finally:
        audit_disabled.reset(token)


def _migrate(xlsm_path: str, session) -> Report:
    rep = Report()
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)

    # 1) Teams (onglet Equipes) ------------------------------------------------
    teams: dict[str, Team] = {}
    for row in wb[SHEET_TEAMS].iter_rows(min_row=2, values_only=True):
        name = _s(row[0])
        if not name:
            continue
        teams[name] = Team(
            name=name,
            manager=_s(row[1]),
            capacite_etp=_num(row[2]),
            description=_s(row[3]),
        )

    # 2) Charges (Détail Mensuel) — dédup par (project_id, team, month), dernier gagnant
    load_rows: dict[tuple[str, str, date], dict] = {}
    referenced_teams: set[str] = set()
    for row in wb[SHEET_LOADS].iter_rows(min_row=2, values_only=True):
        pid, team, month = _s(row[0]), _s(row[2]), _first_of_month(row[3])
        if not (pid and team and month):
            if any(c is not None for c in row[:5]):
                rep.skipped_loads += 1
            continue
        referenced_teams.add(team)
        load_rows[(pid, team, month)] = {
            "project_id": pid,
            "team": team,
            "month": month,
            "days": _num(row[4]) or 0.0,
            "in_plan": _norm_bool(row[6]),
            "updated_at": _to_dt(row[5]),
        }

    # 3) Capacités (Détail Capa Mensuelle) — dédup par (team, month)
    capa_rows: dict[tuple[str, date], dict] = {}
    for row in wb[SHEET_CAPA].iter_rows(min_row=2, values_only=True):
        team, month = _s(row[0]), _first_of_month(row[1])
        if not (team and month):
            if any(c is not None for c in row[:7]):
                rep.skipped_capa += 1
            continue
        referenced_teams.add(team)
        capa_rows[(team, month)] = {
            "team": team,
            "month": month,
            "etp_team": _num(row[2]),
            "etp_projet": _num(row[3]),
            "part_projet": _num(row[4]),
            "jours_indispo": _num(row[5]),
            "capa_projet": _num(row[6]),
            "updated_at": _to_dt(row[7]),
        }

    # 3b) Équipes orphelines → stubs (préserve les FK et les données)
    for team in referenced_teams - set(teams):
        teams[team] = Team(name=team, manager=None, capacite_etp=None,
                            description="(stub — absente de l'onglet Equipes)")
        rep.stub_teams.append(team)

    # 4) Projets (Project Portfolio 25-26, col A→O)
    projects: dict[str, Project] = {}
    for row in wb[SHEET_PROJECTS].iter_rows(min_row=2, values_only=True):
        pid = _s(row[1])
        if not pid:
            continue
        # Lignes fantômes Excel : un project_id traîne mais aucune donnée réelle. project_name
        # est obligatoire (§3.1) → on écarte (évite qu'une ligne vide écrase un vrai projet).
        if _s(row[3]) is None:
            rep.warnings.append(f"ligne projet ignorée (id {pid} sans nom — ligne fantôme)")
            continue
        status = _s(row[5])
        projects[pid] = Project(
            project_id=pid,
            entite=_s(row[0]),
            domain_lead=_s(row[2]),
            project_name=_s(row[3]) or "(sans nom)",
            project_leader=_s(row[4]),
            status=status,
            priorite=_s(row[6]),
            pilier_strategique=_s(row[7]),
            budget_item=_s(row[8]),
            budget_owner=_s(row[9]),
            programme=_s(row[10]),
            in_plan=is_in_plan(status),  # recalculé (§6.2), pas recopié de la col « IN PLAN ? »
            start_date=_to_date(row[12]),
            end_date=_to_date(row[13]),
            last_update=_to_dt(row[14]),  # seed ; recalculé ci-dessous si des charges existent
            total_project_load=None,
        )

    # 4b) Charges orphelines (project_id absent du portefeuille) : signalées et écartées
    orphan_pids = {pid for (pid, _, _) in load_rows} - set(projects)
    if orphan_pids:
        rep.warnings.append(
            f"{len(orphan_pids)} charge(s) ignorée(s) : project_id absent du portefeuille "
            f"({', '.join(sorted(orphan_pids))})"
        )
        load_rows = {k: v for k, v in load_rows.items() if k[0] in projects}

    # 4c) Conformité référentiels §4 — signalée, jamais bloquante (legacy conservé tel quel).
    for field, allowed in _REFERENTIAL_FIELDS.items():
        bad = sorted({
            getattr(p, field) for p in projects.values()
            if getattr(p, field) is not None and getattr(p, field) not in allowed
        })
        if bad:
            rep.warnings.append(
                f"{field}: {len(bad)} valeur(s) hors référentiel §4 : {', '.join(bad)}"
            )

    # 5) Recalculs §3.1 : total_project_load & last_update (§6.6, privilégie les lignes in_plan)
    load_sum: dict[str, float] = defaultdict(float)
    last_upd: dict[str, datetime] = {}
    last_upd_inplan: dict[str, datetime] = {}
    for (pid, _team, _month), v in load_rows.items():
        load_sum[pid] += v["days"] or 0.0
        ts = v["updated_at"]
        if ts:
            if not last_upd.get(pid) or ts > last_upd[pid]:
                last_upd[pid] = ts
            if v["in_plan"] and (not last_upd_inplan.get(pid) or ts > last_upd_inplan[pid]):
                last_upd_inplan[pid] = ts
    for pid, proj in projects.items():
        proj.total_project_load = round(load_sum.get(pid, 0.0), 1)
        recomputed = last_upd_inplan.get(pid) or last_upd.get(pid)
        if recomputed:
            proj.last_update = recomputed  # §6.6

    # 6) Persistance transactionnelle, en **upsert par clé naturelle** (idempotent).
    #    Teams (PK name) et Projects (PK project_id) : merge direct.
    for team in teams.values():
        session.merge(team)
    session.flush()
    for proj in projects.values():
        session.merge(proj)
    session.flush()

    #    Loads/Capacity : PK surrogate → upsert sur la clé d'unicité métier.
    existing_loads = {
        (l.project_id, l.team, l.month): l for l in session.query(MonthlyLoad).all()
    }
    for key, v in load_rows.items():
        cur = existing_loads.get(key)
        if cur is None:
            session.add(MonthlyLoad(**v))
        else:
            for f, val in v.items():
                setattr(cur, f, val)

    existing_capa = {
        (c.team, c.month): c for c in session.query(MonthlyCapacity).all()
    }
    for key, v in capa_rows.items():
        cur = existing_capa.get(key)
        if cur is None:
            session.add(MonthlyCapacity(**v))
        else:
            for f, val in v.items():
                setattr(cur, f, val)

    session.commit()

    rep.teams = len(teams)
    rep.projects = len(projects)
    rep.loads = len(load_rows)
    rep.capacity = len(capa_rows)
    return rep


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Migration xlsm → base (Lot 1, §9).")
    parser.add_argument("xlsm", help="Chemin du classeur source .xlsm")
    parser.add_argument("--database-url", default=settings.database_url,
                        help="Override de DATABASE_URL (défaut : config app).")
    parser.add_argument("--reset", action="store_true",
                        help="Vide les 4 tables avant import (DROP + CREATE).")
    args = parser.parse_args(argv)

    # L'import massif initial n'est pas journalisé (§8.2 concerne les écritures applicatives).
    from app.audit import audit_disabled
    audit_disabled.set(True)

    engine = create_engine(args.database_url, future=True)
    if args.reset:
        Base.metadata.drop_all(engine)
    # Crée les tables manquantes (pratique en dev ; en prod, lancer `alembic upgrade head` avant).
    Base.metadata.create_all(engine)

    Session = sessionmaker(bind=engine, future=True, expire_on_commit=False)
    with Session() as session:
        try:
            rep = migrate(args.xlsm, session)
        except Exception:
            session.rollback()
            raise
    print(rep.render())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
